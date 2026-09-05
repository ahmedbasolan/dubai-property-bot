"""Export investment reports to PDF and Excel."""

import sys
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any

sys.path.insert(0, str(Path(__file__).parent))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
)

from rag import structured_search
from calculators import calculate_mortgage, calculate_str
from developer_scorecard import score_developer, get_all_developer_scores


# --- Styles ---
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

REC_FILLS = {
    "INVEST": GREEN_FILL,
    "HOLD": YELLOW_FILL,
    "AVOID": RED_FILL,
}


def generate_excel_report(
    output_path: str,
    mortgage_rate: float = 4.5,
    tenure: int = 25,
    down_payment_pct: float = 20,
) -> str:
    """Generate a full Excel investment report."""
    result = structured_search("", filters=None)
    scores = result["community_scores"]
    transactions = result["transactions"][:20]

    wb = Workbook()

    # --- Sheet 1: Community Scores ---
    ws = wb.active
    ws.title = "Community Scores"
    headers = [
        "Community", "District", "Composite Score", "Recommendation",
        "Net Yield %", "Gross Yield %", "Price/sqft (AED)", "Avg Price (AED)",
        "Service Charge", "Supply Risk", "Pipeline %", "Occupancy %",
        "Developer", "Price Grade", "Yield Grade", "Net Yield Grade",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    for row_idx, cs in enumerate(scores, 2):
        data = [
            cs["community"], cs["district"], cs["composite_score"], cs["recommendation"],
            cs["avg_net_yield_pct"], cs["avg_roi_pct"], cs["avg_price_per_sqft"],
            cs["avg_price"], cs["avg_service_charge"], cs["supply_risk"],
            cs["pipeline_pct_of_stock"], cs["occupancy_rate"], cs["master_developer"],
            cs["price_score"], cs["yield_score"], cs["net_yield_score"],
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = BORDER
            if col == 4:  # Recommendation column
                cell.fill = REC_FILLS.get(val, PatternFill())
                cell.font = Font(bold=True)

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # --- Sheet 2: Properties with Mortgage ---
    ws2 = wb.create_sheet("Properties")
    headers2 = [
        "Transaction ID", "Community", "Type", "Bedrooms", "Price (AED)",
        "Size (sqft)", "Price/sqft", "Gross Yield %", "Net Yield %",
        "Floor", "View", "Developer", "Status",
        "Monthly Mortgage", "Total Acquisition", "DLD Fee",
    ]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    for row_idx, tx in enumerate(transactions, 2):
        mortgage = calculate_mortgage(
            property_price=tx["price_aed"],
            down_payment_pct=down_payment_pct / 100,
            interest_rate=mortgage_rate,
            tenure_years=tenure,
            size_sqft=tx["size_sqft"],
            service_charge_sqft=tx.get("service_charge_aed_sqft", 15),
        )
        data = [
            tx["transaction_id"], tx["community"], tx["property_type"],
            tx["bedrooms"], tx["price_aed"], tx["size_sqft"],
            tx.get("price_per_sqft", 0), tx["roi_pct"], tx["net_yield_pct"],
            tx.get("floor_level", ""), tx.get("view_type", ""),
            tx.get("developer", ""), tx.get("handover_status", ""),
            mortgage.monthly_payment, mortgage.total_acquisition_cost,
            mortgage.dld_transfer_fee,
        ]
        for col, val in enumerate(data, 1):
            cell = ws2.cell(row=row_idx, column=col, value=val)
            cell.border = BORDER

    for col in range(1, len(headers2) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    # --- Sheet 3: Developer Scores ---
    ws3 = wb.create_sheet("Developers")
    dev_scores = get_all_developer_scores()
    headers3 = [
        "Developer", "Overall Grade", "On-Time %", "Quality",
        "Post-Handover Appreciation %", "Service Charge Grade",
        "Projects", "Units Delivered", "RERA Compliant", "Recommendation",
    ]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    for row_idx, ds in enumerate(dev_scores, 2):
        data = [
            ds.name, ds.overall_grade, ds.on_time_delivery_pct,
            ds.quality_rating, ds.post_handover_appreciation_pct,
            ds.service_charge_efficiency, ds.project_count,
            ds.total_units_delivered, ds.rera_compliance, ds.recommendation,
        ]
        for col, val in enumerate(data, 1):
            cell = ws3.cell(row=row_idx, column=col, value=val)
            cell.border = BORDER

    for col in range(1, len(headers3) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 22

    # --- Sheet 4: Summary ---
    ws4 = wb.create_sheet("Summary")
    ws4.cell(row=1, column=1, value="Dubai Property Investment Report").font = Font(bold=True, size=14)
    ws4.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws4.cell(row=3, column=1, value=f"Mortgage Rate: {mortgage_rate}% | Tenure: {tenure} years | Down Payment: {down_payment_pct}%")
    ws4.cell(row=5, column=1, value="Metric").font = HEADER_FONT
    ws4.cell(row=5, column=1).fill = HEADER_FILL
    ws4.cell(row=5, column=2, value="Value").font = HEADER_FONT
    ws4.cell(row=5, column=2).fill = HEADER_FILL

    summary_data = [
        ("Total Communities", len(scores)),
        ("Total Transactions", len(transactions)),
        ("INVEST Communities", sum(1 for s in scores if s["recommendation"] == "INVEST")),
        ("HOLD Communities", sum(1 for s in scores if s["recommendation"] == "HOLD")),
        ("AVOID Communities", sum(1 for s in scores if s["recommendation"] == "AVOID")),
        ("Avg Net Yield", f"{sum(s['avg_net_yield_pct'] for s in scores) / len(scores):.2f}%"),
        ("Avg Composite Score", f"{sum(s['composite_score'] for s in scores) / len(scores):.0f}/100"),
    ]
    for i, (metric, value) in enumerate(summary_data, 6):
        ws4.cell(row=i, column=1, value=metric).border = BORDER
        ws4.cell(row=i, column=2, value=value).border = BORDER

    ws4.column_dimensions["A"].width = 25
    ws4.column_dimensions["B"].width = 20

    wb.save(output_path)
    return output_path


def generate_pdf_report(
    output_path: str,
    mortgage_rate: float = 4.5,
    tenure: int = 25,
    down_payment_pct: float = 20,
) -> str:
    """Generate a PDF investment report."""
    doc = SimpleDocTemplate(output_path, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        "CustomTitle", parent=styles["Title"],
        fontSize=18, spaceAfter=6*mm, textColor=colors.HexColor("#1F4E79"),
    )
    elements.append(Paragraph("Dubai Property Investment Report", title_style))
    elements.append(Paragraph(
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | "
        f"Rate: {mortgage_rate}% | Tenure: {tenure}yr | DP: {down_payment_pct}%",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 4*mm))
    elements.append(HRFlowable(width="100%", color=colors.HexColor("#1F4E79")))
    elements.append(Spacer(1, 4*mm))

    # Community Scores
    result = structured_search("", filters=None)
    scores = result["community_scores"]

    elements.append(Paragraph("Investment Leaderboard", styles["Heading2"]))
    elements.append(Spacer(1, 2*mm))

    table_data = [["Community", "Score", "Net Yield", "Price/sqft", "Supply Risk", "Rec"]]
    for cs in scores[:15]:
        table_data.append([
            cs["community"][:25],
            f"{cs['composite_score']:.0f}/100",
            f"{cs['avg_net_yield_pct']:.1f}%",
            f"AED {cs['avg_price_per_sqft']:,.0f}",
            cs["supply_risk"],
            cs["recommendation"],
        ])

    table = Table(table_data, colWidths=[55*mm, 20*mm, 22*mm, 28*mm, 22*mm, 20*mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4F8")]),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 6*mm))

    # Top Properties
    transactions = result["transactions"][:10]
    elements.append(Paragraph("Top Properties by Net Yield", styles["Heading2"]))
    elements.append(Spacer(1, 2*mm))

    prop_data = [["ID", "Community", "Type", "Price", "Yield", "Net Yield", "Mortgage/mo"]]
    for tx in transactions:
        mortgage = calculate_mortgage(
            property_price=tx["price_aed"],
            down_payment_pct=down_payment_pct / 100,
            interest_rate=mortgage_rate,
            tenure_years=tenure,
            size_sqft=tx["size_sqft"],
            service_charge_sqft=tx.get("service_charge_aed_sqft", 15),
        )
        prop_data.append([
            tx["transaction_id"],
            tx["community"][:20],
            f"{tx['bedrooms']}BR",
            f"AED {tx['price_aed']:,}",
            f"{tx['roi_pct']}%",
            f"{tx['net_yield_pct']:.1f}%",
            f"AED {mortgage.monthly_payment:,}",
        ])

    prop_table = Table(prop_data, colWidths=[22*mm, 35*mm, 14*mm, 28*mm, 16*mm, 18*mm, 25*mm])
    prop_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("FONTSIZE", (0, 1), (-1, -1), 6.5),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4F8")]),
    ]))
    elements.append(prop_table)
    elements.append(Spacer(1, 6*mm))

    # Developer Scores
    dev_scores = get_all_developer_scores()
    elements.append(Paragraph("Developer Scorecard", styles["Heading2"]))
    elements.append(Spacer(1, 2*mm))

    dev_data = [["Developer", "Grade", "On-Time", "Quality", "Appreciation", "RERA"]]
    for ds in sorted(dev_scores, key=lambda x: x.overall_grade):
        dev_data.append([
            ds.name,
            ds.overall_grade,
            f"{ds.on_time_delivery_pct:.0f}%",
            ds.quality_rating,
            f"{ds.post_handover_appreciation_pct:.1f}%",
            "✓" if ds.rera_compliance else "✗",
        ])

    dev_table = Table(dev_data, colWidths=[35*mm, 18*mm, 22*mm, 18*mm, 28*mm, 16*mm])
    dev_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4F8")]),
    ]))
    elements.append(dev_table)
    elements.append(Spacer(1, 6*mm))

    # Disclaimer
    elements.append(HRFlowable(width="100%", color=colors.grey))
    elements.append(Spacer(1, 2*mm))
    disclaimer_style = ParagraphStyle(
        "Disclaimer", parent=styles["Normal"],
        fontSize=7, textColor=colors.grey, alignment=1,
    )
    elements.append(Paragraph(
        "DISCLAIMER: This report is for informational purposes only. Not financial advice. "
        "Data is from a mock dataset. Verify with DLD and consult a RERA-licensed advisor.",
        disclaimer_style,
    ))

    doc.build(elements)
    return output_path
